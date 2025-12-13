from typing import Optional
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from model.base_processor import BaseProcessor
from model.data_class.settings import OvertimeSettings
from model.helper.export_file_formatter import ExportFileFormatter, WorkbookType
from model.helper.save_utils import save_target_workbooks
from model.helper.date_utils import format_date, try_parse_date

class OvertimeComparator(BaseProcessor):
    def __init__(self, formatter: Optional[ExportFileFormatter] = None):
        super().__init__()
        self.formatter = formatter or ExportFileFormatter()
        self.overtime_index: dict[str, dict] = {}
        
    def compare(
        self,
        settings: dict,
        date_start_str: str,
        date_end_str: str,
        overtime_file: str,
        hris_file: str
    ) -> None:
        self.overtime_index = {}
        
        overtime_settings = self.apply_overtime_settings(settings)
        source_wb = self.load_source_wb(overtime_file)
        hris_wb = self.load_hris_wb(hris_file)
        output_dir = self.get_output_dir(overtime_file)
        source_ws = self.get_source_sheets(source_wb, overtime_settings.sheet_names)
        hris_ws = self.get_hris_source_sheets(hris_wb)
        
        print(f"Date Start: {date_start_str}, Date End: {date_end_str}")
        
        targets = {
            code: self.formatter.prepare_workbook(code, WorkbookType.COMPARE)
            for code, checked in overtime_settings.company_codes.items()
            if checked
        }

        for ws in source_ws:
            self._process_overtime_sheet(ws, overtime_settings, targets, date_start_str, date_end_str)
        for ws in hris_ws:
            self._process_hris_sheet(ws, date_start_str, date_end_str)
        
        self._print_overtime_index(self.overtime_index, targets)
            
        save_target_workbooks(
            targets=targets,
            output_dir=output_dir,
            date_start_str=date_start_str,
            date_end_str=date_end_str,
            type_str="Overtime Comparison",
            template_name=settings.get("template_name"),
            formatter=self.formatter,
        )
        
    def _process_overtime_sheet(
        self, 
        ws: Worksheet,
        settings: OvertimeSettings,
        targets: dict[str, Workbook],
        date_start_str: str, 
        date_end_str: str
    ) -> None:
        # Determine company code by ws title
        ws_title = ws.title
        sheet_company_code = self._company_code_from_sheet_title(ws_title)
        print(f"Processing sheet: {ws.title} for company code: {sheet_company_code}")
            
        if not sheet_company_code:
            return
        if sheet_company_code not in targets:
            return
            
        # Find last non-empty cell in row_counter_col, from bottom up
        last_data_row = settings.data_start_row
        for row in range(ws.max_row, settings.data_start_row - 1, -1):
            value = ws.cell(row=row, column=settings.row_counter_col).value
            if value is not None and str(value).strip() != "":
                last_data_row = row
                break
            
        # Initialize persistent variables
        employee_id = ""
        employee_name = ""
        notes = ""
            
        for row in range(settings.data_start_row, last_data_row + 1):
                
            date = ws.cell(row=row, column=settings.date_col).value
            shift = ws.cell(row=row, column=settings.shift_col).value
            overtime = ws.cell(row=row, column=settings.ovt_col).value
            overtime_hours = ws.cell(row=row, column=settings.ovt_hour_col).value
                
            _id = ws.cell(row=row, column=settings.employee_id_col).value
            _name = ws.cell(row=row, column=settings.employee_name_col).value
            _notes = ws.cell(row=row, column=settings.notes_col).value
                
            # Parse date as a date object and compare ranges using dates
            formatted_date = format_date(date)
            parsed_date = try_parse_date(formatted_date)
            if parsed_date is None:
                continue
                
            try:
                start_dt = try_parse_date(date_start_str) or parsed_date
                end_dt = try_parse_date(date_end_str) or parsed_date
            except Exception:
                start_dt = parsed_date
                end_dt = parsed_date

            if not (start_dt <= parsed_date <= end_dt):
                continue
                
            if not shift or not overtime or not overtime_hours:
                continue
                
            # Update persistent variables if current row has new values
            employee_id = str(_id).strip() if _id else employee_id
            employee_name = str(_name).strip() if _name else employee_name
            notes = str(_notes).strip() if _notes else notes
                
            status, timein, timeout = self.map_status_by_shift(shift)
            
            key = f"{formatted_date}_{employee_id}"
            if key in self.overtime_index:
                self.overtime_index[key]["overtime"] += overtime
                continue
            
            self.overtime_index[key] = {
                "hris_overtime": 0,
                "date": formatted_date,
                "employee_id": employee_id,
                "employee_name": employee_name,
                "status": status,
                "overtime": overtime,
                "time_in": timein,
                "time_out": timeout,
                "notes": notes,
                "company_code": sheet_company_code
            }
                
    def _process_hris_sheet(
        self, 
        ws: Worksheet,
        date_start_str: str, 
        date_end_str: str
    ) -> None:
        print(f"Processing HRIS sheet: {ws.title}")
        
        start_row = 2
        start_date_col = 8
        id_col = 3
        name_col = 2
        
        # Extract dates from header row
        header_dates = []
        for col in range(start_date_col, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            try:
                date = format_date(cell_value)
                header_dates.append((col, date))
            except Exception:
                continue
            
        # Determine start and end columns based on date range
        start_col = None
        end_col = None
        
        for col, date in header_dates:
            if date == date_start_str and start_col is None:
                start_col = col
            if date == date_end_str:
                end_col = col

        if start_col is None:
            start_col = start_date_col
        if end_col is None:
            end_col = ws.max_column
            
        print(f"Data rows: {start_row} to {ws.max_row}, Columns: {start_col} to {end_col}")
        
        for row in range(start_row, ws.max_row + 1):
            employee_id_raw = ws.cell(row=row, column=id_col).value
            if not employee_id_raw:
                continue
            employee_id = str(employee_id_raw).strip()
            
            for col in range(start_col, end_col + 1):
                overtime = ws.cell(row=row, column=col).value
                date = ws.cell(row=1, column=col).value
                
                if not date:
                    continue
                if not overtime or str(overtime).strip() == "":
                    overtime = 0
                    
                formatted_date = format_date(date)
                employee_id_str = str(employee_id).strip()
                key = f"{formatted_date}_{employee_id_str}"
                matched_overtime_record = self.overtime_index.get(key)
                
                if matched_overtime_record:
                    matched_overtime_record["hris_overtime"] = float(overtime)

    def _print_overtime_index(self, overtime_index: dict[str, dict], targets: dict[str, Workbook]):
        for key, record in overtime_index.items():
            target_ws = targets[record["company_code"]].active
            target_ws.append([
                record["overtime"],
                record["hris_overtime"],
                record["overtime"]-record["hris_overtime"],
                record["date"],
                record["employee_id"],
                record["employee_name"],
                record["status"],
                record["overtime"],
                record["time_in"],
                record["time_out"],
                record["notes"]
            ])        
                    
                    
                
                
                

              
    