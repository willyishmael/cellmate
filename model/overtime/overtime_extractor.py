from typing import Optional
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from model.base_processor import BaseProcessor
from model.data_class.settings import OvertimeSettings
from model.helper.export_file_formatter import ExportFileFormatter, WorkbookType
from model.helper.save_utils import save_target_workbooks
from model.helper.date_utils import format_date, try_parse_date

class OvertimeExtractor(BaseProcessor):
    def __init__(
        self,
        formatter: Optional[ExportFileFormatter]= None
    ):
        super().__init__()
        self.formatter = formatter or ExportFileFormatter()
        
    def extract(
        self,
        settings: dict,
        date_start_str: str,
        date_end_str: str, 
        overtime_file: str  
    ):
        print(f"OvertimeExtractor: Starting extraction for file: {overtime_file}")
        overtime_settings = self.apply_overtime_settings(settings)
        source_wb = self.load_source_wb(overtime_file)
        output_dir = self.get_output_dir(overtime_file)
        source_ws = self.get_source_sheets(source_wb, overtime_settings.sheet_names)
        print(f"Date Start: {date_start_str}, Date End: {date_end_str}")
        
        targets = {
            code: self.formatter.prepare_workbook(code, WorkbookType.EXTRACT)
            for code, checked in overtime_settings.company_codes.items()
            if checked
        }

        overtime_index = self._process_source_sheet(source_ws, overtime_settings, targets, date_start_str, date_end_str)
        self._print_overtime_index(overtime_index, targets)
        self._save_output_files(targets, output_dir, date_start_str, date_end_str)
        
        save_target_workbooks(
            targets=targets,
            output_dir=output_dir,
            date_start_str=date_start_str,
            date_end_str=date_end_str,
            formatter=self.formatter,
        )
        
    def _process_source_sheet(
        self, 
        source_ws: list[Worksheet],
        settings: OvertimeSettings,
        target_ws: dict[str, Workbook],
        date_start_str: str, 
        date_end_str: str
    ) -> dict[str, dict]:
        overtime_index: dict[str, dict] = {}
        for ws in source_ws:
            # Determine company code by ws title
            ws_title = ws.title
            company_code = self._company_code_from_sheet_title(ws_title)
            print(f"Processing sheet: {ws.title} for company code: {company_code}")
            
            if not company_code:
                return
            if company_code not in target_ws:
                return
            
            # Find last non-empty cell in row_counter_col, from bottom up
            last_data_row = settings.data_start_row
            for row in range(ws.max_row, settings.data_start_row - 1, -1):
                value = ws.cell(row=row, column=settings.row_counter_col).value
                if value is not None and str(value).strip() != "":
                    last_data_row = row
                    break
            
            # Initialize persistent variables
            employee_id = ""
            employee_name = ""
            notes = ""
            
            for row in range(settings.data_start_row, last_data_row + 1):
                
                date = ws.cell(row=row, column=settings.date_col).value
                shift = ws.cell(row=row, column=settings.shift_col).value
                overtime = ws.cell(row=row, column=settings.ovt_col).value
                overtime_hours = ws.cell(row=row, column=settings.ovt_hour_col).value
                
                _id = str(ws.cell(row=row, column=settings.employee_id_col).value).strip()
                _name = str(ws.cell(row=row, column=settings.employee_name_col).value).strip()
                _notes = str(ws.cell(row=row, column=settings.notes_col).value).strip()
                
                # Parse date as a date object and compare ranges using dates
                formatted_date = format_date(date)
                parsed_date = try_parse_date(formatted_date)
                if parsed_date is None:
                    continue
                
                try:
                    start_dt = try_parse_date(date_start_str) or parsed_date
                    end_dt = try_parse_date(date_end_str) or parsed_date
                except Exception:
                    start_dt = parsed_date
                    end_dt = parsed_date
                
                # Update persistent variables if current row has new values
                none = (None, "", "None")
                employee_id = _id if _id not in none else employee_id
                employee_name = _name if _name not in none else employee_name
                notes = _notes if _notes not in none else notes
                status, timein, timeout = self.map_status(shift=shift)

                # Skip rows outside date range or with invalid data
                if not (start_dt <= parsed_date <= end_dt):
                    continue
                
                if not shift or not overtime or not overtime_hours:
                    continue

                key = f"{formatted_date}_{employee_id}"
                if key in overtime_index:
                    overtime_index[key]["overtime"] += overtime
                    continue
                    
                overtime_index[key] = {
                    "date": formatted_date,
                    "employee_id": employee_id,
                    "employee_name": employee_name,
                    "status": status,
                    "overtime": overtime,
                    "timein": timein,
                    "timeout": timeout,
                    "notes": notes,
                    "company_code": company_code
                }
                
        return overtime_index
            
    def _print_overtime_index(self, overtime_index: dict[str, dict], targets: dict[str, Workbook]):
        for key, record in overtime_index.items():
            target_ws = targets[record["company_code"]].active
            target_ws.append([
                record["date"],
                record["employee_id"],
                record["employee_name"],
                record["status"],
                record["overtime"],
                record["timein"],
                record["timeout"],
                record["notes"]
            ])
            
        
        
        