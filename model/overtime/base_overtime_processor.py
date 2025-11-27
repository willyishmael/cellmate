from datetime import datetime, date
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import re


class BaseOvertimeProcessor():
    def __init__(self) -> None:
        self.settings = {}
        self.overtime_wb = None
        self.hris_wb = None
        self.sheet_names = []
        self.company_codes = {}
        
    # Public API
    def apply_settings(self, settings: dict) -> None:
        """
        Apply or update settings before running processing.
        Should be called every time before extract/compare.
        """
        if not isinstance(settings, dict):
            raise ValueError("Settings must be a dictionary.")
        self.settings = settings
        self._normalize_settings()
        
    # Settings Normalization
    def _normalize_settings(self) -> None:
        """Normalize numeric and list-based settings for consistent use."""
        s = self.settings
        
        def to_int(value, default):
            try:
                return int(value)
            except Exception:
                return default
            
        self.employee_id_col = to_int(s.get("employee_id_column"), 2)
        self.data_start_row = to_int(s.get("data_start_row"), 5)
        self.row_counter_col = to_int(s.get("row_counter_column"), 1)
        
        self.sheet_names = self._parse_comma_list(s.get("sheet_names", ""))
        self.company_codes = s.get("company_codes", {})
        
    def _parse_comma_list(self, value: str) -> list[str]:
        """Convert comma-separated text into a list of trimmed strings."""
        if not value:
            return []
        return [x.strip() for x in value.split(",") if x.strip()]
    
    def load_overtime_wb(self, file_path: str) -> Workbook:
        """Load Overtime Excel file."""
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {path}")
        self.overtime_wb = load_workbook(path, data_only=True)
        return self.overtime_wb
    
    def load_hris_wb(self, hris_file: str) -> Workbook:
        """Load HRIS Excel file."""
        path = Path(hris_file)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {path}")
        self.hris_wb = load_workbook(path, data_only=True)
        return self.hris_wb
    
    def get_output_dir(self, file_path: str) -> Path:
        """Return directory where output files should be saved."""
        return Path(file_path).parent
    
    def get_overtime_source_sheet(self) -> list[Worksheet]:
        """Return sheet objects based on settings.sheet_names."""
        if not self.overtime_wb:
            raise ValueError("Overtime workbook is not loaded.")
        
        sheets = []
        for name in self.sheet_names:
            if name in self.overtime_wb.sheetnames:
                sheets.append(self.overtime_wb[name])
        return sheets
    
    def get_hris_source_sheet(self) -> list[Worksheet]:
        """Return all sheet objects."""
        if not self.hris_wb:
            raise ValueError("HRIS workbook is not loaded.")
        return [self.hris_wb[s] for s in self.hris_wb.sheetnames]

    def _company_code_from_sheet_title(self, title: str) -> str | None:
        """Determine company code from a worksheet title.

        Looks for known company code substrings in the sheet title (case-insensitive).
        Order matters: check longer codes first (e.g. 'PTM' before 'PM').
        Returns the matched code string (e.g. 'PTM', 'TMP', 'PM') or None if no match.
        """
        if not title:
            return None
        t = str(title).upper()
        for code in ("PTM", "TMP", "PM"):
            if code in t:
                return code
        return None
    
    def map_status(self, shift: str) -> tuple[str, str, str]:
        """Map attendance codes to descriptions and time ranges."""
        code = str(shift).strip().upper()
        mapping = {
            "PAGI": ("Hadir (H)", "07:00", "18:00"),
            "SIANG": ("Hadir (H)", "07:00", "18:00"),
            "MALAM": ("Hadir shift malam (HM)", "19:00", "06:00"),
        }
        return mapping.get(code, (code, "", ""))