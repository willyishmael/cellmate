from openpyxl import load_workbook, Workbook
from datetime import datetime
from pathlib import Path

class AttendanceExtract:
    def __init__(self):
        self.settings = {}
        self.wb = None
        self.sheet_names = []
        self.ignore_list = []
        self.data = []
    
    def extract(self, settings: dict, date_start_str: str, date_end_str: str, file: str):
        self.settings = settings
        self.wb = load_workbook(file, data_only=True)
        source_path = Path(file)
        output_dir = source_path.parent
        
        sheets_str = settings.get("sheet_names", "")
        self.sheet_names = [
            emp.strip() for emp in sheets_str.split(",") if emp.strip()
         ]
        
        ignore_str = settings.get("ignore_list", "")
        self.ignore_list = [
            emp.strip() for emp in ignore_str.split(",") if emp.strip()
         ]
        
        ws_sources = [self.wb[s] for s in self.sheet_names if s in self.wb.sheetnames]
        
        company_codes = settings.get("company_codes", {})
        targets = {
            code: self._init_target_sheet(code)
            for code, checked in company_codes.items()
            if checked
        }
        
        for ws in ws_sources:
            self._process_source_sheet(ws, targets, date_start_str, date_end_str)
            
        for code, twb in targets.items():
            out_path = output_dir / f"{code}_output.xlsx"
            twb.save(out_path)
            
    def _init_target_sheet(self, company_code):
        wb = Workbook()
        ws = wb.active
        ws.title = company_code
        headers = ["Tanggal", "Employee ID", "Nama Karyawan", "Status", 
                   "Overtime", "Time In", "Time Out", "Keterangan"]
        ws.append(headers)
        return wb
    
    def _process_source_sheet(self, ws, targets, date_start_str, date_end_str):
        s = self.settings
        start_row = int(s["data_start_row"])
        id_col = int(s["employee_id_column"])
        name_col = int(s["employee_name_column"])
        company_code_col = int(s["company_code_column"])
        header_row = int(s["date_header_row"])
        row_counter_col = int(s["row_counter_column"])

        # VBA-like logic: find last non-empty cell in row_counter_col, from bottom up
        last_data_row = start_row
        for row in range(ws.max_row, start_row - 1, -1):
            value = ws.cell(row=row, column=row_counter_col).value
            if value is not None and str(value).strip() != "":
                last_data_row = row
                break
        
        # Extract dates from header row
        header_dates = []
        for col in range(company_code_col + 1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col).value
            try:
                date = self._format_date(cell_value)
                header_dates.append((col, date))
            except Exception:
                continue
            
        # Determine start and end columns based on date range
        start_col = None
        end_col = None
        
        for col, date in header_dates:
            print(f"col {col}, date {date}")
            if date == date_start_str and start_col is None:
                start_col = col
            if date == date_end_str:
                end_col = col

        if start_col is None:
            start_col = company_code_col + 1
        if end_col is None:
            end_col = ws.max_column
        
        print(f"Start column str: {date_start_str}, End column str: {date_end_str}")
        print(f"Start column: {start_col}, End column: {end_col}")
        
        # Process each row of data
        for row in range(start_row, last_data_row + 1):
            employee_id = ws.cell(row=row, column=id_col).value
            employee_name = ws.cell(row=row, column=name_col).value
            company_code = ws.cell(row=row, column=company_code_col).value

            if not employee_id or str(employee_id).strip() in self.ignore_list:
                continue

            if company_code not in targets:
                continue

            for col in range(start_col, end_col + 1):
                code = ws.cell(row=row, column=col).value
                date = ws.cell(row=header_row, column=col).value

                if not code or str(code).strip() == "" or not date:
                    continue

                ws_target = targets[company_code].active

                status, timein, timeout = self._map_status(code)

                ws_target.append([
                    self._format_date(date),
                    employee_id,
                    employee_name if employee_name else "",
                    status,
                    0,  # Overtime
                    timein,
                    timeout,
                    ""   # Keterangan
                ])
                
    def _map_status(self, code: str):
        """Map code to status and time in/out like VBA Select Case."""
        code = str(code).strip().upper()
        mapping = {
            "H":  ("Hadir (H)", "07:00", "18:00"),
            "HM": ("Hadir shift malam (HM)", "19:00", "06:00"),
            "A":  ("Alpa (A)", "", ""),
            "S":  ("Sakit (S)", "", ""),
            "I":  ("Izin (I)", "", ""),
            "HC": ("Cuti (C)", "", ""),
            "DLK":("Dinas Luar Kota (DLK)", "", ""),
            "OFF":("OFF", "", ""),
        }
        return mapping.get(code, (code, "", ""))
    
    def _format_date(self, value):
        """Ensure date is formatted as yyyy-mm-dd string."""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        try:
            return datetime.strptime(str(value), "%Y-%m-%d").strftime("%Y-%m-%d")
        except Exception:
            return str(value)
    
       