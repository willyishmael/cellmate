from enum import Enum
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

class WorkbookType(Enum):
    EXTRACT = "extract"
    COMPARE = "compare"

class ExportFileFormatter:
    
    def prepare_workbook(self, company_code: str, type: WorkbookType) -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = company_code
        header = []
        
        extract_headers = ["Tanggal", "Employee ID", "Nama Karyawan", "Status",
                           "Overtime", "Time In", "Time Out", "Keterangan"]
        compare_headers = ["Manual", "HRIS", "Difference"] + extract_headers
        
        if type == WorkbookType.EXTRACT:
            header = extract_headers
        elif type == WorkbookType.COMPARE:
            header = compare_headers
        else:
            raise ValueError(f"Unknown type for workbook preparation: {type}")
        
        ws.append(header)
        return wb
    
    def format_worksheet(self, ws: Worksheet):
        # Auto-fit columns
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except Exception:
                    pass
            ws.column_dimensions[column].width = max_length + 2
        
        # Header styling
        sheet_name = ws.title.upper()
        color = {
            "PTM": "FFCCE5FF",
            "TMP": "FFFFE5CC",
            "PM":  "FFCCFFCC",
        }
        fill_color = color.get(sheet_name, "FFFFFFFF")
        
        for cell in ws[1]:
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.font = Font(bold=True)
            
        # Add auto-filter to header row
        ws.auto_filter.ref = ws.dimensions