import logging
from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from model.helper.export_file_formatter import ExportFileFormatter

logger = logging.getLogger(__name__)


def copy_values_only(src_wb: Workbook) -> Workbook:
    """Return a new Workbook containing only cell values (no formulas, no relationships)."""
    new_wb = Workbook()
    # remove default created sheet
    try:
        default = new_wb.active
        new_wb.remove(default)
    except Exception:
        pass

    for sheet in src_wb.worksheets:
        tgt: Worksheet = new_wb.create_sheet(title=sheet.title)
        for row in sheet.iter_rows(values_only=True):
            # row is a tuple of values (not cells)
            tgt.append(list(row))
    return new_wb


def save_workbook_with_fallback(src_wb: Workbook, out_path: Path, formatter: ExportFileFormatter =None):
    """Try to save workbook; on failure attempt to clear externalReferences, then fallback to value-only copy.

    formatter: optional object with method `format_worksheet(ws)`; if provided, it will be called
    on the workbook's active sheet before saving.
    """
    try:
        if formatter is not None:
            try:
                formatter.format_worksheet(src_wb.active)
            except Exception:
                logger.exception("Formatter failed on source workbook; continuing to save attempt")

        src_wb.save(out_path)
        logger.info("Saved %s", out_path)
        return
    except Exception as exc:
        logger.exception("Initial save failed for %s", out_path)

    # Try to clear externalReferences metadata on package if present
    try:
        pkg = getattr(src_wb, "package", None)
        if pkg is not None and getattr(pkg, "externalReferences", None):
            try:
                pkg.externalReferences = []
                if formatter is not None:
                    try:
                        formatter.format_worksheet(src_wb.active)
                    except Exception:
                        logger.exception("Formatter failed on source workbook after clearing externalReferences")
                src_wb.save(out_path)
                logger.info("Saved %s after clearing externalReferences", out_path)
                return
            except Exception:
                logger.exception("Save after clearing externalReferences also failed for %s", out_path)
    except Exception:
        logger.exception("Error while attempting to inspect/clear package.externalReferences")

    # Final fallback: copy values only and save that workbook
    try:
        new_wb = copy_values_only(src_wb)
        if formatter is not None:
            try:
                formatter.format_worksheet(new_wb.active)
            except Exception:
                logger.exception("Formatter failed on value-only workbook; will still attempt save")
        new_wb.save(out_path)
        logger.warning("Saved value-only copy to %s (formulas/links removed)", out_path)
        return
    except Exception:
        logger.exception("Final fallback (value-only copy) failed for %s", out_path)
        raise


def save_target_workbooks(
    targets: dict,
    output_dir: Path,
    date_start_str: str,
    date_end_str: str,
    type_str: str,
    template_name: str,
    formatter: ExportFileFormatter | None = None,
    filename_suffix: str = ".xlsx",
    name_template_single: str = "{date} {template} {code} {type} {suffix}",
    name_template_range: str = "{start} to {end} {template} {code} {type} {suffix}",
):
    """Save multiple target workbooks to `output_dir`.

    - Ensures `output_dir` exists (creates if necessary).
    - Builds filenames using provided templates.
    - Uses `save_workbook_with_fallback` for robust saving.
    """
    try:
        output_dir.mkdir(parents=True, exist_ok=True)
    except Exception:
        logger.exception("Failed to create output directory %s", output_dir)
        raise

    for code, twb in targets.items():
        try:
            if date_end_str == date_start_str:
                file_name = name_template_single.format(
                    date=date_start_str, 
                    template=template_name, 
                    code=code, 
                    type=type_str, 
                    suffix=filename_suffix
                )
            else:
                file_name = name_template_range.format(
                    start=date_start_str, 
                    end=date_end_str,
                    template=template_name, 
                    code=code, 
                    type=type_str, 
                    suffix=filename_suffix
                )

            out_path = output_dir / file_name
            if formatter is not None:
                try:
                    formatter.format_worksheet(twb.active)
                except Exception:
                    logger.exception("Formatter failed for workbook %s; continuing to save attempt", file_name)

            save_workbook_with_fallback(twb, out_path, formatter=formatter)
            logger.info("Saved %s", out_path)
        except Exception:
            logger.exception("Failed to save workbook for code %s", code)
