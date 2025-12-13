from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from model.data_class.settings import AttendanceSettings, OvertimeOptDrvSettings, OvertimeSettings

class BaseProcessor:
    
    def apply_attendance_settings(self, settings: dict[str, any]) -> AttendanceSettings:
        """
        Apply or update settings before running processing.
        Should be called every time before extract/compare.
        Returns an AttendanceSettings object with parsed configuration.
        """

        self.attendance_settings = AttendanceSettings(
            employee_id_col=int(settings.get("employee_id_column") or 2),
            employee_name_col=int(settings.get("employee_name_column") or 3),
            company_code_col=int(settings.get("company_code_column") or 5),
            data_start_row=int(settings.get("data_start_row") or 5),
            date_header_row=int(settings.get("date_header_row") or 4),
            row_counter_col=int(settings.get("row_counter_column") or 1),
            sheet_names=self._parse_comma_list(settings.get("sheet_names", "")),
            ignore_list=self._parse_comma_list(settings.get("ignore_list", "")),
            company_codes=settings.get("company_codes", {}),
            time_off_only=settings.get("time_off_only", False)
        )

        return self.attendance_settings
    
    def apply_overtime_settings(self, settings: dict[str, any]) -> OvertimeSettings:
        """
        Apply or update settings before running processing.
        Should be called every time before extract/compare.
        Returns an OvertimeSettings object with parsed configuration.
        """

        # parse primary columns and derive related column indices
        emp_col = int(settings.get("employee_id_column") or 2)
        data_start = int(settings.get("data_start_row") or 5)
        row_counter = int(settings.get("row_counter_column") or 1)

        name_col = emp_col + 1
        date_col = name_col + 2
        shift_col = date_col + 1
        ovt_start_col = shift_col + 1
        ovt_end_col = ovt_start_col + 1
        ovt_hour_col = ovt_end_col + 1
        ovt_col = ovt_hour_col + 1
        notes_col = ovt_col + 2

        self.overtime_settings = OvertimeSettings(
            employee_id_col=emp_col,
            data_start_row=data_start,
            row_counter_col=row_counter,
            sheet_names=self._parse_comma_list(settings.get("sheet_names", "")),
            company_codes=settings.get("company_codes", {}),
            employee_name_col=name_col,
            date_col=date_col,
            shift_col=shift_col,
            ovt_start_col=ovt_start_col,
            ovt_end_col=ovt_end_col,
            ovt_hour_col=ovt_hour_col,
            ovt_col=ovt_col,
            notes_col=notes_col
        )

        return self.overtime_settings
    
    def apply_overtime_optdrv_settings(self, settings: dict[str, any]) -> OvertimeOptDrvSettings:
        """
        Apply or update settings before running processing.
        Should be called every time before extract/compare.
        Returns an OvertimeOptDrvSettings object with parsed configuration.
        """

        self.overtime_optdrv_settings = OvertimeOptDrvSettings(
            employee_id_col=int(settings.get("employee_id_column") or 2),
            data_start_row=int(settings.get("data_start_row") or 5),
            employee_name_col=int(settings.get("employee_name_column") or 3),
            date_header_row=int(settings.get("date_header_row") or 4),
            row_counter_col=int(settings.get("row_counter_column") or 1),
            company_code_col=int(settings.get("company_code_column") or 5),
            sheet_names=self._parse_comma_list(settings.get("sheet_names", "")),
            company_codes=settings.get("company_codes", {})
        )

        return self.overtime_optdrv_settings
    
    def load_source_wb(self, file_path: str) -> Workbook:
        """Load Source Excel file. This will handle both attendance and overtime files."""
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {path}")
        source_wb = load_workbook(path, data_only=True)
        return source_wb
    
    def load_hris_wb(self, hris_file: str) -> Workbook:
        """Load HRIS Excel file. This will handle both attendance and overtime HRIS files."""
        path = Path(hris_file)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {path}")
        hris_wb = load_workbook(path, data_only=True)
        return hris_wb

    def get_output_dir(self, file_path: str) -> Path:
        """Return directory where output files should be saved."""
        return Path(file_path).parent

    def get_source_sheets(self, source_wb: Workbook, sheet_names: list[str]) -> list[Worksheet]:
        """Return sheet objects based on settings.sheet_names."""
        if not source_wb:
            raise ValueError("Workbook not loaded yet.")
        if not sheet_names:
            return [source_wb.active]
        return [source_wb[sheet] for sheet in sheet_names if sheet in source_wb.sheetnames]
    
    def get_hris_source_sheets(self, hris_wb: Workbook) -> list[Worksheet]:
        """Return sheet objects based on settings.sheet_names."""
        if not hris_wb: 
            raise ValueError("Workbook not loaded yet.")
        return [hris_wb[sheet] for sheet in hris_wb.sheetnames]
    
    def map_status_by_code(self, code: str) -> tuple[str, str, str]:
        """Map attendance codes to descriptions and time ranges."""
        code = str(code).strip().upper()
        mapping = {
            "H":  ("Hadir (H)", "07:00", "18:00"),
            "HM": ("Hadir shift malam (HM)", "19:00", "06:00"),
            "A":  ("Alpa (A)", "", ""),
            "S":  ("Sakit (S)", "", ""),
            "I":  ("Izin (I)", "", ""),
            "HC": ("Cuti (C)", "", ""),
            "DLK":("Dinas Luar Kota (DLK)", "", ""),
            "OFF":("OFF", "", ""),
        }
        return mapping.get(code, ("", "", ""))
    
    def map_status_by_shift(self, shift: str) -> tuple[str, str, str]:
        """Map shift to descriptions and time ranges."""
        code = str(shift).strip().upper()
        mapping = {
            "PAGI": ("Hadir (H)", "07:00", "18:00"),
            "SIANG": ("Hadir (H)", "07:00", "18:00"),
            "MALAM": ("Hadir shift malam (HM)", "19:00", "06:00"),
        }
        return mapping.get(code, (code, "", ""))
    
    def _parse_comma_list(self, value: str) -> list[str]:
        """Convert comma-separated text into a list of trimmed strings."""
        if not value:
            return []
        return [x.strip() for x in value.split(",") if x.strip()]
    
    def _company_code_from_sheet_title(self, title: str) -> str | None:
        """Determine company code from a worksheet title."""
        if not title:
            return None
        t = str(title).upper()
        for code in ("PTM", "TMP", "PM"):
            if code in t:
                return code
        return None