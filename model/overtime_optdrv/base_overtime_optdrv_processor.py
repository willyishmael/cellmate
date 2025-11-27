from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class BaseOvertimeOptdrvProcessor:
    """
    Base class for overtime optdrv processing (shared between extraction and comparison).
    Handles normalization, workbook loading, and utility methods.
    """

    def __init__(self) -> None:
        self.settings = {}
        self.overtime_wb = None
        self.hris_wb = None
        self.sheet_names = []
        self.ignore_list = []
        self.company_codes = {}

    # Public API
    def apply_settings(self, settings: dict):
        """
        Apply or update settings before running processing.
        Should be called every time before extract/compare.
        """
        if not isinstance(settings, dict):
            raise ValueError("Settings must be a dictionary.")
        self.settings = settings
        self._normalize_settings()

    # Settings Normalization
    def _normalize_settings(self):
        """Normalize numeric and list-based settings for consistent use."""
        s = self.settings

        def to_int(value, default):
            try:
                return int(value)
            except Exception:
                return default

        self.employee_id_col = to_int(s.get("employee_id_column"), 2)
        self.employee_name_col = to_int(s.get("employee_name_column"), 3)
        self.company_code_col = to_int(s.get("company_code_column"), 5)
        self.data_start_row = to_int(s.get("data_start_row"), 5)
        self.date_header_row = to_int(s.get("date_header_row"), 4)
        self.row_counter_col = to_int(s.get("row_counter_column"), 1)
        self.sheet_names = self._parse_comma_list(s.get("sheet_names", ""))
        self.company_codes = s.get("company_codes", {})

    @staticmethod
    def _parse_comma_list(value: str):
        """Convert comma-separated text into a list of trimmed strings."""
        if not value:
            return []
        return [x.strip() for x in value.split(",") if x.strip()]

    def load_overtime_wb(self, file_path: str) -> Workbook:
        """Load Overtime Excel file."""
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {path}")
        self.overtime_wb = load_workbook(path, data_only=True)
        return self.overtime_wb
    
    def load_hris_wb(self, hris_file: str) -> Workbook:
        """Load HRIS Excel file."""
        path = Path(hris_file)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {path}")
        self.hris_wb = load_workbook(path, data_only=True)
        return self.hris_wb
    
    def get_output_dir(self, file_path: str) -> Path:
        """Return directory where output files should be saved."""
        return Path(file_path).parent
    
    def get_overtime_source_sheet(self) -> list[Worksheet]:
        """Return sheet objects based on settings.sheet_names."""
        if not self.overtime_wb:
            raise ValueError("Overtime workbook is not loaded.")
        
        sheets = []
        for name in self.sheet_names:
            if name in self.overtime_wb.sheetnames:
                sheets.append(self.overtime_wb[name])
        return sheets
    
    def get_hris_source_sheet(self) -> list[Worksheet]:
        """Return all sheet objects."""
        if not self.hris_wb:
            raise ValueError("HRIS workbook is not loaded.")
        return [self.hris_wb[s] for s in self.hris_wb.sheetnames]
    
    
    