from typing import Optional
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from model.base_processor import BaseProcessor
from model.data_class.settings import OvertimeOptDrvSettings
from model.helper.date_utils import format_date
from model.helper.export_file_formatter import ExportFileFormatter, WorkbookType
from model.helper.save_utils import save_target_workbooks


class OvertimeOptdrvExtractor(BaseProcessor):
    def __init__(self, formatter: Optional[ExportFileFormatter] = None):
        super().__init__()
        self.formatter = formatter or ExportFileFormatter()
        
    def extract(
        self,
        settings: dict,
        date_start_str: str,
        date_end_str: str, 
        overtime_file: str  
    ):
        print(f"OvertimeOptdrvExtractor: Starting extraction for file: {overtime_file}")
        overtime_settings = self.apply_overtime_optdrv_settings(settings)
        source_wb = self.load_source_wb(overtime_file)
        output_dir = self.get_output_dir(overtime_file)
        source_ws = self.get_source_sheets(source_wb, overtime_settings.sheet_names)
        
        print(f"Date Start: {date_start_str}, Date End: {date_end_str}")
        
        targets = {
            code: self.formatter.prepare_workbook(code, WorkbookType.EXTRACT)
            for code, checked in overtime_settings.company_codes.items()
            if checked
        }

        for ws in source_ws:
            self._process_source_sheet(ws, targets, date_start_str, date_end_str)
            
        save_target_workbooks(
            targets=targets,
            output_dir=output_dir,
            date_start_str=date_start_str,
            date_end_str=date_end_str,
            type_str="Overtime Optdrv",
            template_name=settings.get("template_name"),
            formatter=self.formatter,
        )
    
    def _process_source_sheet(
        self, 
        ws: Worksheet,
        settings: OvertimeOptDrvSettings,
        targets: dict[str, Workbook], 
        date_start_str: str, 
        date_end_str: str
    ) -> None:
        """Process a single source worksheet and populate target workbooks."""
        print(f"Processing source sheet: {ws.title}")
        
        # Find last non-empty cell in row_counter_col, from bottom up
        last_data_row = settings.data_start_row
        for row in range(ws.max_row, settings.data_start_row - 1, -1):
            value = ws.cell(row=row, column=settings.row_counter_col).value
            if value is not None and str(value).strip() != "":
                last_data_row = row
                break
        
        # Extract dates from header row
        header_dates = []
        for col in range(settings.company_code_col + 1, ws.max_column + 1):
            cell_value = ws.cell(row=settings.date_header_row, column=col).value
            try:
                date = format_date(cell_value)
                header_dates.append((col, date))
            except Exception:
                continue
            
        # Determine start and end columns based on date range
        start_col = None
        end_col = None
        
        for col, date in header_dates:
            if date == date_start_str and start_col is None:
                start_col = col
            if date == date_end_str:
                end_col = col

        if start_col is None:
            start_col = settings.company_code_col + 1
        if end_col is None:
            end_col = ws.max_column
        
        for row in range(settings.data_start_row, last_data_row + 1):
            company_code = ws.cell(row=row, column=settings.company_code_col).value
            
            if not company_code or company_code not in targets:
                continue
            
            target_wb = targets[company_code]
            target_ws = target_wb.active
            
            employee_id = ws.cell(row=row, column=settings.employee_id_col).value
            employee_name = ws.cell(row=row, column=settings.employee_name_col).value
            
            for col in range(start_col, end_col + 1):
                date = None
                for hd_col, hd_date in header_dates:
                    if hd_col == col:
                        date = hd_date
                        break
                if not date:
                    continue
                
                overtime = ws.cell(row=row, column=col).value
                if overtime is None or str(overtime).strip() == "" or overtime == 0:
                    continue
                
                time_in = "07:00"
                time_out = "19:00"
                status = "Hadir (H)"
                
                target_ws.append([
                    date,
                    employee_id,
                    employee_name,
                    status,
                    overtime,
                    time_in,
                    time_out
                ])
        
