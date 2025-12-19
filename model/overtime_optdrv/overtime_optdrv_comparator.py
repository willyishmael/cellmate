from typing import Optional
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from model.base_processor import BaseProcessor
from model.data_class.settings import OvertimeOptDrvSettings
from model.helper.date_utils import format_date
from model.helper.export_file_formatter import ExportFileFormatter, WorkbookType
from model.helper.save_utils import save_target_workbooks


class OvertimeOptdrvComparator(BaseProcessor):
    def __init__(self, formatter: Optional[ExportFileFormatter] = None):
        super().__init__()
        self.formatter = formatter or ExportFileFormatter()
        self.overtime_index: dict[str, dict] = {}
        
    def compare(
        self,
        settings: dict,
        date_start_str: str,
        date_end_str: str,
        overtime_file: str,
        hris_file: str
    ) -> None:
        self.overtime_index = {}
        
        overtime_settings = self.apply_overtime_optdrv_settings(settings)
        source_wb = self.load_source_wb(overtime_file)
        hris_wb = self.load_hris_wb(hris_file)
        output_dir = self.get_output_dir(overtime_file)
        source_ws = self.get_source_sheets(source_wb, overtime_settings.sheet_names)
        hris_ws = self.get_hris_source_sheets(hris_wb)
        print(f"Date Start: {date_start_str}, Date End: {date_end_str}")
        
        targets = {
            code: self.formatter.prepare_workbook(code, WorkbookType.COMPARE)
            for code, checked in overtime_settings.company_codes.items()
            if checked
        }
        
        for ws in source_ws:
            self._process_overtime_sheet(ws, overtime_settings, targets, date_start_str, date_end_str)
        for ws in hris_ws:
            self._process_hris_sheet(ws, overtime_settings, targets, date_start_str, date_end_str)
        
        self._print_overtime_index(self.overtime_index, targets)
        
        save_target_workbooks(
            targets=targets,
            output_dir=output_dir,
            date_start_str=date_start_str,
            date_end_str=date_end_str,
            type_str="Overtime Optdrv Comparison",
            template_name=settings.get("template_name"),
            formatter=self.formatter,
        )
        
    def _process_overtime_sheet(
        self, 
        ws: Worksheet,
        settings: OvertimeOptDrvSettings,
        targets: dict[str, Workbook], 
        date_start_str: str, 
        date_end_str: str
    ) -> None:
        print(f"Processing source sheet: {ws.title}")
        
        # Find last non-empty cell in row_counter_col, from bottom up
        last_data_row = settings.data_start_row
        for row in range(ws.max_row, settings.data_start_row - 1, -1):
            value = ws.cell(row=row, column=settings.row_counter_col).value
            if value is not None and str(value).strip() != "":
                last_data_row = row
                break
        
        # Extract dates from header row
        header_dates = []
        for col in range(settings.company_code_col + 1, ws.max_column + 1):
            cell_value = ws.cell(row=settings.date_header_row, column=col).value
            try:
                date = format_date(cell_value)
                header_dates.append((col, date))
            except Exception:
                continue
            
        # Determine start and end columns based on date range
        start_col = None
        end_col = None
        
        for col, date in header_dates:
            if date == date_start_str and start_col is None:
                start_col = col
            if date == date_end_str:
                end_col = col

        if start_col is None:
            start_col = settings.company_code_col + 1
        if end_col is None:
            end_col = ws.max_column
            
        for row in range(settings.data_start_row, last_data_row + 1):
            company_code = ws.cell(row=row, column=settings.company_code_col).value
            if not company_code or company_code not in targets:
                continue
            
            employee_id = ws.cell(row=row, column=settings.employee_id_col).value
            employee_name = ws.cell(row=row, column=settings.employee_name_col).value
            if not employee_id:
                continue
            
            for col in range(start_col, end_col + 1):
                date = ws.cell(row=settings.date_header_row, column=col).value
                overtime = ws.cell(row=row, column=col).value
                
                if not date:
                    continue
                if overtime in (None, "", " "):
                    continue
                
                time_in = "07:00"
                time_out = "19:00"
                status = "Hadir (H)"
                
                formatted_date = format_date(date)
                key = f"{formatted_date}_{employee_id}"
                record = {
                    "hris_overtime": 0,
                    "date": formatted_date,
                    "employee_id": employee_id,
                    "employee_name": employee_name,
                    "status": status,
                    "overtime": float(overtime),
                    "time_in": time_in,
                    "time_out": time_out,
                    "notes": "",
                    "company_code": company_code
                }
            
                if key in self.overtime_index:
                    self.overtime_index[key]["overtime"] += float(overtime)
                else:
                    self.overtime_index[key] = record
                
    def _process_hris_sheet(
        self, 
        ws: Worksheet,
        settings: OvertimeOptDrvSettings,
        targets: dict[str, Workbook],
        date_start_str: str, 
        date_end_str: str
    ) -> None:
        print(f"Processing HRIS sheet: {ws.title}")
        
        start_row = 2
        start_date_col = 8
        id_col = 3
        name_col = 2
        
        # Extract dates from header row
        header_dates = []
        for col in range(start_date_col, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            try:
                date = format_date(cell_value)
                header_dates.append((col, date))
            except Exception:
                continue
            
        # Determine start and end columns based on date range
        start_col = None
        end_col = None
        
        for col, date in header_dates:
            if date == date_start_str and start_col is None:
                start_col = col
            if date == date_end_str:
                end_col = col

        if start_col is None:
            start_col = start_date_col
        if end_col is None:
            end_col = ws.max_column
            
        print(f"Data rows: {start_row} to {ws.max_row}, Columns: {start_col} to {end_col}")
        
        for row in range(start_row, ws.max_row + 1):
            employee_id_raw = ws.cell(row=row, column=id_col).value
            if not employee_id_raw:
                continue
            employee_id = str(employee_id_raw).strip()
            
            for col in range(start_col, end_col + 1):
                overtime = ws.cell(row=row, column=col).value
                date = ws.cell(row=1, column=col).value
                
                if not date:
                    continue
                if not overtime or str(overtime).strip() == "":
                    overtime = 0
                    
                formatted_date = format_date(date)
                employee_id_str = str(employee_id).strip()
                key = f"{formatted_date}_{employee_id_str}"
                matched_overtime_record = self.overtime_index.get(key)

                if matched_overtime_record:
                    matched_overtime_record["hris_overtime"] += float(overtime)
    
    def _print_overtime_index(self, overtime_index: dict[str, dict], targets: dict[str, Workbook]):
        for key, record in overtime_index.items():
            target_ws = targets[record["company_code"]].active
            target_ws.append([
                record["overtime"],
                record["hris_overtime"],
                record["overtime"]-record["hris_overtime"],
                record["date"],
                record["employee_id"],
                record["employee_name"],
                record["status"],
                record["overtime"],
                record["time_in"],
                record["time_out"],
                record["notes"]
            ])   
    