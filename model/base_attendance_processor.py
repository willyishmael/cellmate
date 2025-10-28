from openpyxl import load_workbook
from datetime import datetime
from pathlib import Path
import re


class BaseAttendanceProcessor:
    """
    Base class for attendance processing (shared between extraction and comparison).
    Handles normalization, workbook loading, and utility methods.
    """

    def __init__(self):
        self.settings = {}
        self.attendance_wb = None
        self.hris_wb = None
        self.sheet_names = []
        self.ignore_list = []
        self.company_codes = {}

    # Public API
    def apply_settings(self, settings: dict):
        """
        Apply or update settings before running processing.
        Should be called every time before extract/compare.
        """
        if not isinstance(settings, dict):
            raise ValueError("Settings must be a dictionary.")
        self.settings = settings
        self._normalize_settings()

    # Settings Normalization
    def _normalize_settings(self):
        """Normalize numeric and list-based settings for consistent use."""
        s = self.settings

        def to_int(value, default):
            try:
                return int(value)
            except Exception:
                return default

        self.employee_id_col = to_int(s.get("employee_id_column"), 2)
        self.employee_name_col = to_int(s.get("employee_name_column"), 3)
        self.company_code_col = to_int(s.get("company_code_column"), 5)
        self.data_start_row = to_int(s.get("data_start_row"), 5)
        self.date_header_row = to_int(s.get("date_header_row"), 4)
        self.row_counter_col = to_int(s.get("row_counter_column"), 1)

        self.sheet_names = self._parse_comma_list(s.get("sheet_names", ""))
        self.ignore_list = self._parse_comma_list(s.get("ignore_list", ""))
        self.company_codes = s.get("company_codes", {})

    @staticmethod
    def _parse_comma_list(value: str):
        """Convert comma-separated text into a list of trimmed strings."""
        if not value:
            return []
        return [x.strip() for x in value.split(",") if x.strip()]

    # File Handling
    def load_attendance_wb(self, file_path: str):
        """Load Attendance Excel file."""
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {path}")
        self.attendance_wb = load_workbook(path, data_only=True)
        return self.attendance_wb
    
    def load_hris_wb(self, hris_file: str):
        """Load HRIS Excel file."""
        path = Path(hris_file)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {path}")
        self.hris_wb = load_workbook(path, data_only=True)
        return self.hris_wb

    def get_output_dir(self, file_path: str):
        """Return directory where output files should be saved."""
        return Path(file_path).parent

    def get_attendance_source_sheets(self):
        """Return sheet objects based on settings.sheet_names."""
        if not self.attendance_wb:
            raise ValueError("Workbook not loaded yet.")
        if not self.sheet_names:
            return [self.attendance_wb.active]
        return [self.attendance_wb[s] for s in self.sheet_names if s in self.attendance_wb.sheetnames]
    
    def get_hris_source_sheets(self):
        """Return sheet objects based on settings.sheet_names."""
        if not self.hris_wb:
            raise ValueError("Workbook not loaded yet.")
        return [self.hris_wb[s] for s in self.hris_wb.sheetnames]

    # Utility Methods
    def _format_date(self, value):
        """Ensure date formatted as yyyy-mm-dd string.

        Handles:
        - datetime objects
        - ISO-like 'YYYY-MM-DD' or 'YYYY/MM/DD'
        - 'DD/MM/YYYY' and 'DD-MM-YYYY'
        - 'DD/MM' (no year) -> assumes current year

        Falls back to returning the original string if parsing fails.
        """
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")

        s = str(value).strip()

        # Common full-date formats
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except Exception:
                pass

        # Handle day/month without year, e.g. '08/10' -> assume current year
        try:
            if re.match(r"^\d{1,2}/\d{1,2}$", s):
                day_str, month_str = s.split("/")
                year = datetime.now().year
                dt = datetime(year=year, month=int(month_str), day=int(day_str))
                return dt.strftime("%Y-%m-%d")
        except Exception:
            pass

        # Fallback: return original string
        return s

    def _map_status(self, code: str):
        """Map attendance codes to descriptions and time ranges."""
        code = str(code).strip().upper()
        mapping = {
            "H":  ("Hadir (H)", "07:00", "18:00"),
            "HM": ("Hadir shift malam (HM)", "19:00", "06:00"),
            "A":  ("Alpa (A)", "", ""),
            "S":  ("Sakit (S)", "", ""),
            "I":  ("Izin (I)", "", ""),
            "HC": ("Cuti (C)", "", ""),
            "DLK":("Dinas Luar Kota (DLK)", "", ""),
            "OFF":("OFF", "", ""),
        }
        return mapping.get(code, (code, "", ""))