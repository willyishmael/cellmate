from typing import Optional
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from model.base_processor import BaseProcessor
from model.data_class.settings import AttendanceSettings
from model.helper.date_utils import format_date
from model.helper.export_file_formatter import ExportFileFormatter, WorkbookType
from model.helper.save_utils import save_target_workbooks

class AttendanceExtractor(BaseProcessor):
    """Class to handle attendance extraction from Excel files."""
    def __init__(self, formatter: Optional[ExportFileFormatter] = None):
        super().__init__()
        self.formatter = formatter or ExportFileFormatter()
    
    def extract(
        self, 
        settings: dict, 
        date_start_str: str, 
        date_end_str: str, 
        file: str
    ) -> None:
        """Run the extraction process with given settings and file."""
        print("Starting extraction process...")
        attendance_settings = self.apply_attendance_settings(settings)
        source_wb = self.load_source_wb(file)
        output_dir = self.get_output_dir(file)
        source_ws = self.get_source_sheets(source_wb, attendance_settings.sheet_names)
        
        # Prepare target workbooks for each selected company code
        targets = {
            code: self.formatter.prepare_workbook(code, WorkbookType.EXTRACT)
            for code, checked in attendance_settings.company_codes.items()
            if checked
        }
        
        for ws in source_ws:
            self._process_source_sheet(ws, attendance_settings, targets, date_start_str, date_end_str)
            
        save_target_workbooks(
            targets=targets,
            output_dir=output_dir,
            date_start_str=date_start_str,
            date_end_str=date_end_str,
            type_str="Attendance",
            formatter=self.formatter,
        )
    
    def _process_source_sheet(
        self, 
        ws: Worksheet,
        settings: AttendanceSettings,
        targets: dict[str, Workbook], 
        date_start_str: str, 
        date_end_str: str
    ) -> None:
        print(f"Processing sheet: {ws.title}")

        # Find last non-empty cell in row_counter_col, from bottom up
        last_data_row = settings.data_start_row
        for row in range(ws.max_row, settings.data_start_row - 1, -1):
            value = ws.cell(row=row, column=settings.row_counter_col).value
            if value is not None and str(value).strip() != "":
                last_data_row = row
                break
        
        # Extract dates from header row
        header_dates = []
        for col in range(settings.company_code_col + 1, ws.max_column + 1):
            cell_value = ws.cell(row=settings.date_header_row, column=col).value
            try:
                date = format_date(cell_value)
                header_dates.append((col, date))
            except Exception:
                continue
            
        # Determine start and end columns based on date range
        start_col = None
        end_col = None
        
        for col, date in header_dates:
            if date == date_start_str and start_col is None:
                start_col = col
            if date == date_end_str:
                end_col = col

        if start_col is None:
            start_col = settings.company_code_col + 1
        if end_col is None:
            end_col = ws.max_column
            
        print(f"Data rows: {settings.data_start_row} to {last_data_row}, Columns: {start_col} to {end_col}")
        
        # Process each row of data
        for row in range(settings.data_start_row, last_data_row + 1):
            employee_id_raw = ws.cell(row=row, column=settings.employee_id_col).value
            if not employee_id_raw:
                continue
            
            employee_id = str(employee_id_raw).strip()
            if employee_id in settings.ignore_list:
                continue
            
            employee_name = ws.cell(row=row, column=settings.employee_name_col).value.strip()
            
            company_code_raw = ws.cell(row=row, column=settings.company_code_col).value
            company_code = str(company_code_raw).strip()
            if not company_code or company_code not in targets:
                continue

            for col in range(start_col, end_col + 1):
                code = ws.cell(row=row, column=col).value
                date_raw = ws.cell(row=settings.date_header_row, column=col).value
                date = ws.cell(row=settings.date_header_row, column=col).value
                if not code or date_raw is None :
                    continue

                formatted_date = format_date(date_raw)
                status, timein, timeout = self.map_status_by_code(code)
                ws_target = targets[company_code].active
                ws_target.append([
                    formatted_date,
                    employee_id,
                    employee_name or "",
                    status,
                    0,  # Overtime
                    timein,
                    timeout,
                    ""   # Keterangan
                ])