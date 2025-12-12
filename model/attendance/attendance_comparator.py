from typing import Optional
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from model.base_processor import BaseProcessor
from model.data_class.settings import AttendanceSettings
from model.helper.date_utils import format_date
from model.helper.export_file_formatter import ExportFileFormatter, WorkbookType
from model.helper.save_utils import save_target_workbooks

class AttendanceComparator(BaseProcessor):
    """
    Compare Attendance data (from extracted report) with HRIS export.
    Outputs differences such as missing uploads or mismatched attendance codes.
    """
    
    def __init__(self, formatter: Optional[ExportFileFormatter] = None):
        super().__init__()
        self.formatter = formatter or ExportFileFormatter()
        self.attendance_index = {} # key -> record
        self.duplicates = [] # optional list to collect duplicates
    
    def compare(
        self, 
        settings: dict, 
        date_start_str: str, 
        date_end_str: str, 
        attendance_file: str,
        hris_file: str
    ) -> None:
        """Run the comparison process with given settings and files."""
        print("Starting comparison process...")
        attendance_settings = self.apply_attendance_settings(settings)
        source_wb = self.load_source_wb(attendance_file)
        hris_wb = self.load_hris_wb(hris_file)
        output_dir = self.get_output_dir(attendance_file)
        source_ws = self.get_source_sheets(source_wb, attendance_settings.sheet_names)
        hris_ws = self.get_hris_source_sheets(hris_wb)
        
        print("Preparing target workbooks...")
        
        # Prepare target workbooks for each selected company code
        targets = {
            code: self.formatter.prepare_workbook(code, WorkbookType.COMPARE)
            for code, checked in attendance_settings.company_codes.items()
            if checked
        }

        print(f"Target workbooks prepared. Company codes: {list(targets.keys())}")
        # Process attendance and HRIS sheets
        for ws in source_ws:
            self._process_attendance_sheet(ws, attendance_settings, targets, date_start_str, date_end_str)
        for ws in hris_ws:
            self._process_hris_sheet(ws, attendance_settings, targets, date_start_str, date_end_str)
            
        save_target_workbooks(
            targets=targets,
            output_dir=output_dir,
            date_start_str=date_start_str,
            date_end_str=date_end_str,
            type_str="Attendance Comparison",
            template_name=settings.get("template_name"),
            formatter=self.formatter,
        )
    
    def _process_attendance_sheet(
        self, 
        ws: Worksheet, 
        settings: AttendanceSettings,
        targets: dict[str, Workbook],
        date_start_str: str, 
        date_end_str: str
    ) -> None:
        print(f"Processing attendance sheet: {ws.title}")
        
        # Find last non-empty cell in row_counter_col, from bottom up
        last_data_row = settings.data_start_row
        for row in range(ws.max_row, settings.data_start_row - 1, -1):
            value = ws.cell(row=row, column=settings.row_counter_col).value
            if value is not None and str(value).strip() != "":
                last_data_row = row
                break
        
        # Extract dates from header row
        header_dates = []
        for col in range(settings.company_code_col + 1, ws.max_column + 1):
            cell_value = ws.cell(row=settings.date_header_row, column=col).value
            try:
                date = format_date(cell_value)
                header_dates.append((col, date))
            except Exception:
                continue
            
        # Determine start and end columns based on date range
        start_col = None
        end_col = None
        
        for col, date in header_dates:
            if date == date_start_str and start_col is None:
                start_col = col
            if date == date_end_str:
                end_col = col

        if start_col is None:
            start_col = settings.company_code_col + 1
        if end_col is None:
            end_col = ws.max_column
            
        print(f"Data rows: {settings.data_start_row} to {last_data_row}, Columns: {start_col} to {end_col}")
        
        # Process each row of data
        for row in range(settings.data_start_row, last_data_row + 1):
            employee_id = ws.cell(row=row, column=settings.employee_id_col).value
            employee_name = ws.cell(row=row, column=settings.employee_name_col).value
            company_code = ws.cell(row=row, column=settings.company_code_col).value

            if not employee_id or str(employee_id).strip() in settings.ignore_list:
                continue

            if company_code not in targets:
                continue

            for col in range(start_col, end_col + 1):
                code = ws.cell(row=row, column=col).value
                date = ws.cell(row=settings.date_header_row, column=col).value
                if not code or str(code).strip() == "" or not date:
                    continue

                formatted_date = format_date(date)
                employee_id_str = str(employee_id).strip()

                key = f"{formatted_date}_{employee_id_str}"
                status, timein, timeout = self.map_status_by_code(code)
                
                record = {
                    "date": formatted_date,
                    "employee_id": employee_id_str,
                    "employee_name": employee_name or "",
                    "company_code": company_code,
                    "status_code": code,
                    "status": status,
                    "overtime": 0,
                    "timein": timein,
                    "timeout": timeout,
                    "notes": ""
                }
                
                if key in self.attendance_index:
                    self.duplicates.append(record)
                else:
                    self.attendance_index[key] = record

                
    def _process_hris_sheet(
        self, 
        ws: Worksheet,
        settings: AttendanceSettings,
        targets: dict[str, Workbook],
        date_start_str: str, 
        date_end_str: str
    ) -> None:
        print(f"Processing HRIS sheet: {ws.title}")
        
        start_row = 2
        id_col = 1
        name_col = 2
        
        # Extract dates from header row
        header_dates = []
        for col in range(5, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            try:
                date = format_date(cell_value)
                header_dates.append((col, date))
            except Exception:
                continue
            
        # Determine start and end columns based on date range
        start_col = None
        end_col = None
        
        for col, date in header_dates:
            if date == date_start_str and start_col is None:
                start_col = col
            if date == date_end_str:
                end_col = col

        if start_col is None:
            start_col = settings.company_code_col + 1
        if end_col is None:
            end_col = ws.max_column
        
        # Process each row of data
        for row in range(start_row, ws.max_row + 1):
            employee_id = ws.cell(row=row, column=id_col).value

            for col in range(start_col, end_col + 1):
                date = ws.cell(row=1, column=col).value
                status = ws.cell(row=row, column=col).value
                if not status or str(status).strip() == "" or not date:
                    continue
                
                formatted_date = format_date(date)
                employee_id_str = str(employee_id).strip()
                key = f"{formatted_date}_{employee_id_str}"
                matched_record = self.attendance_index.get(key)
                
                if matched_record:
                    self._build_attendance_comparison_row(matched_record, status, targets)

    def _build_attendance_comparison_row(self, matched_record: dict, hris_status, targets: dict) -> None:
        """Build and append an attendance comparison row to the appropriate worksheet."""
        company_code = matched_record.get("company_code")
        ws_target = targets.get(company_code).active
        difference = matched_record.get("status") == hris_status
        ws_target.append([
            matched_record.get("status"),  # Manual
            hris_status,  # HRIS
            difference,  # Difference
            matched_record.get("date"),
            matched_record.get("employee_id"),
            matched_record.get("employee_name"),
            matched_record.get("status"),
            matched_record.get("overtime"),
            matched_record.get("timein"),
            matched_record.get("timeout"),
            matched_record.get("notes"),
        ])

        