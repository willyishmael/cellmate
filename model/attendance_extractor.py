from openpyxl import Workbook
from model.helper.export_file_formatter import ExportFileFormatter
from model.base_attendance_processor import BaseAttendanceProcessor

class AttendanceExtractor(BaseAttendanceProcessor):
    """Class to handle attendance extraction from Excel files."""
    def __init__(self):
        super().__init__()
        self.formatter = ExportFileFormatter()
    
    def extract(self, settings: dict, date_start_str: str, date_end_str: str, file: str):
        """Run the extraction process with given settings and file."""
        print("Starting extraction process...")
        self.apply_settings(settings)
        self.load_attendance_wb(file)
        output_dir = self.get_output_dir(file)
        ws_sources = self.get_attendance_source_sheets()
        
        # Prepare target workbooks for each selected company code
        targets = {
            code: self._init_target_sheet(code)
            for code, checked in self.company_codes.items()
            if checked
        }
        
        for ws in ws_sources:
            self._process_source_sheet(ws, targets, date_start_str, date_end_str)
        
        # Save output files
        print("Saving output files...")
        print(f"Targets Items: {list(targets.keys())}")
        for code, twb in targets.items():
            print(f"Saving output for company code: {code}")
            file_name = (
                f"{date_start_str} {code} Attendance.xlsx"
                if date_end_str == date_start_str
                else f"{date_start_str} to {date_end_str} {code} Attendance.xlsx"
            )
            out_path = output_dir / file_name
            self.formatter.format_worksheet(twb.active)
            twb.save(out_path)
            print(f"Saved {out_path.name}")
            
    # Internal Helpers
    def _init_target_sheet(self, company_code):
        wb = Workbook()
        ws = wb.active
        ws.title = company_code
        headers = ["Tanggal", "Employee ID", "Nama Karyawan", "Status", 
                   "Overtime", "Time In", "Time Out", "Keterangan"]
        ws.append(headers)
        return wb
    
    def _process_source_sheet(self, ws, targets, date_start_str, date_end_str):
        print(f"Processing sheet: {ws.title}")
        
        start_row = self.data_start_row
        id_col = self.employee_id_col
        name_col = self.employee_name_col
        company_code_col = self.company_code_col
        header_row = self.date_header_row
        row_counter_col = self.row_counter_col

        # Find last non-empty cell in row_counter_col, from bottom up
        last_data_row = start_row
        for row in range(ws.max_row, start_row - 1, -1):
            value = ws.cell(row=row, column=row_counter_col).value
            if value is not None and str(value).strip() != "":
                last_data_row = row
                break
        
        # Extract dates from header row
        header_dates = []
        for col in range(company_code_col + 1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col).value
            try:
                date = self._format_date(cell_value)
                header_dates.append((col, date))
            except Exception:
                continue
            
        # Determine start and end columns based on date range
        start_col = None
        end_col = None
        
        for col, date in header_dates:
            if date == date_start_str and start_col is None:
                start_col = col
            if date == date_end_str:
                end_col = col

        if start_col is None:
            start_col = company_code_col + 1
        if end_col is None:
            end_col = ws.max_column
            
        print(f"Data rows: {start_row} to {last_data_row}, Columns: {start_col} to {end_col}")
        
        # Process each row of data
        for row in range(start_row, last_data_row + 1):
            employee_id = ws.cell(row=row, column=id_col).value
            employee_name = ws.cell(row=row, column=name_col).value
            company_code = ws.cell(row=row, column=company_code_col).value

            if not employee_id or str(employee_id).strip() in self.ignore_list:
                continue

            if company_code not in targets:
                continue

            for col in range(start_col, end_col + 1):
                code = ws.cell(row=row, column=col).value
                date = ws.cell(row=header_row, column=col).value
                if not code or str(code).strip() == "" or not date:
                    continue

                ws_target = targets[company_code].active
                status, timein, timeout = self._map_status(code)
                ws_target.append([
                    self._format_date(date),
                    employee_id,
                    employee_name if employee_name else "",
                    status,
                    0,  # Overtime
                    timein,
                    timeout,
                    ""   # Keterangan
                ])